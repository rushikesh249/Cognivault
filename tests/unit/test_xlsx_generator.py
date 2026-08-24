"""Unit tests for XLSX Spreadsheet Report Generator (TRD Section 22, Component #17)."""

import os
from pathlib import Path
import openpyxl
import pytest

from backend.app.documents.templates.spreadsheet_report import (
    render_spreadsheet_report,
    sanitize_cell_value,
)


@pytest.fixture
def sample_audit_data():
    return {
        "title": "PRIMARY FLARE HEADER & PIPING INTEGRITY AUDIT",
        "task_id": "TASK-XLSX-TEST-01",
        "facility": "Olefins Cracker Unit 3 - Flare Line",
        "timestamp": "2026-08-24T18:00:00Z",
        "status": "ACTION REQUIRED — NON-COMPLIANCE DETECTED",
        "summary": "Autonomous audit executed against API 570 piping inspection standards and safety manual clauses.",
        "critical_findings": [
            "Flange bolt wall thinning exceeds 1.5mm critical threshold.",
            "Pressure relief valve PRV-204 calibration overdue by 60 days.",
            "Drain valve seal gland weeping hydrocarbon traces.",
        ],
        "compliance_gaps": [
            ("Flange Bolt Thinning", "Safety SOP - Section 4.2", "CRITICAL NON-COMPLIANCE"),
            ("PRV-204 Calibration", "Equipment Standards - Section 11.4", "MAJOR GAP"),
            ("Seal Gland Weeping", "Maintenance Manual - Section 8.1", "MODERATE GAP"),
        ],
        "recommendations": [
            "Schedule immediate depressurization and flange bolt replacement.",
            "Perform bench calibration on PRV-204 within 48 hours.",
            "Repack mechanical seal gland during next turnaround window.",
        ],
    }


def test_xlsx_generation_success(sample_audit_data, tmp_path):
    """Verify XLSX report renders successfully with all 4 expected sheets and content."""
    output_path = tmp_path / "report.xlsx"
    res = render_spreadsheet_report(sample_audit_data, output_path)

    assert res.exists()
    assert res.stat().st_size > 1000

    # Load workbook and verify sheets
    wb = openpyxl.load_workbook(str(output_path), data_only=True)
    expected_sheets = ["Summary & Metadata", "Inspection Findings", "Compliance Gaps & Citations", "Recommendations"]
    assert wb.sheetnames == expected_sheets

    # Verify Summary Sheet
    ws_sum = wb["Summary & Metadata"]
    assert "PRIMARY FLARE HEADER" in str(ws_sum["A3"].value)
    assert "TASK-XLSX-TEST-01" in str(ws_sum["B5"].value)
    assert "ACTION REQUIRED" in str(ws_sum["B8"].value)

    # Verify Findings Sheet
    ws_fnd = wb["Inspection Findings"]
    assert ws_fnd.max_row >= 8  # header + findings
    assert "FND-01" in str(ws_fnd["A6"].value)
    assert "Flange bolt wall thinning" in str(ws_fnd["B6"].value)

    # Verify Compliance Sheet
    ws_comp = wb["Compliance Gaps & Citations"]
    assert "GAP-01" in str(ws_comp["A6"].value)
    assert "Safety SOP - Section 4.2" in str(ws_comp["C6"].value)

    # Verify Recommendations Sheet
    ws_recs = wb["Recommendations"]
    assert "P1" in str(ws_recs["A6"].value)
    assert "depressurization" in str(ws_recs["B6"].value)

    wb.close()


def test_xlsx_formula_injection_sanitization():
    """Verify formula injection triggers (=, +, -, @, \\t, \\r) are neutralized (CWE-1236)."""
    dangerous_inputs = [
        "=SUM(A1:A10)",
        "+1+1",
        "-5*5",
        "@SUM(B1:B5)",
        "\tCMD|' /C calc'!A0",
        "\r=1+1",
    ]
    for raw in dangerous_inputs:
        sanitized = sanitize_cell_value(raw)
        assert sanitized.startswith("'"), f"Dangerous input '{raw}' was not escaped!"

    # Safe inputs must not be modified
    safe_inputs = [
        "Normal finding text",
        "Section 4.2 Relief valve",
        "TASK-123",
        12345,
        0.95,
    ]
    for safe in safe_inputs:
        assert sanitize_cell_value(safe) == safe


def test_xlsx_disclaimer_presence_on_all_worksheets(sample_audit_data, tmp_path):
    """Verify mandatory non-certified draft disclaimer is present on all worksheets."""
    output_path = tmp_path / "disclaimer_check.xlsx"
    render_spreadsheet_report(sample_audit_data, output_path)

    wb = openpyxl.load_workbook(str(output_path), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cell_a1 = str(ws["A1"].value or "")
        assert "AI-Generated Draft" in cell_a1, f"Missing disclaimer in sheet '{sheet_name}'"
        assert "Non-Certified Verdict" in cell_a1, f"Missing non-certified notice in sheet '{sheet_name}'"
    wb.close()
