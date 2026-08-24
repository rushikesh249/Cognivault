"""XLSX Spreadsheet Report Generator (TRD Section 22, Component #17)."""

import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Palette & Fonts
FONT_NAME = "Segoe UI"
NAVY_FILL = PatternFill(start_color="102C57", end_color="102C57", fill_type="solid")
ICE_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F9FBFC", end_color="F9FBFC", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="102C57")
SECTION_FONT = Font(name=FONT_NAME, size=13, bold=True, color="102C57")
SUBSECTION_FONT = Font(name=FONT_NAME, size=11, bold=True, color="333333")
BOLD_FONT = Font(name=FONT_NAME, size=10, bold=True, color="222222")
REGULAR_FONT = Font(name=FONT_NAME, size=10, color="333333")
ALERT_FONT = Font(name=FONT_NAME, size=10, bold=True, color="B40000")
DISCLAIMER_FONT = Font(name=FONT_NAME, size=9, italic=True, bold=True, color="B40000")

THIN_SIDE = Side(border_style="thin", color="D0D7DE")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_BORDER = Border(
    left=Side(border_style="thin", color="102C57"),
    right=Side(border_style="thin", color="102C57"),
    top=Side(border_style="medium", color="102C57"),
    bottom=Side(border_style="medium", color="102C57"),
)


def sanitize_cell_value(val: Any) -> Any:
    """
    Sanitize user-controlled values to prevent CSV/Excel Formula Injection (CWE-1236).
    Neutralizes values beginning with '=', '+', '-', '@', '\t', or '\r'.
    """
    if isinstance(val, str):
        if val.startswith(("=", "+", "-", "@", "\t", "\r")):
            return f"'{val}"
    return val


def auto_fit_columns(ws, min_width: int = 12, max_width: int = 60):
    """Dynamically adjust column widths based on content length with sensible bounds."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            lines = val_str.split("\n")
            for line in lines:
                if len(line) > max_len and len(line) < 120:
                    max_len = len(line)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


def render_spreadsheet_report(data: Dict[str, Any], output_path: Path) -> Path:
    """
    Render professional 4-sheet XLSX technical audit workbook from structured data.

    Worksheets:
    1. Summary & Metadata
    2. Inspection Findings
    3. Compliance & SOP Citations
    4. Action Recommendations
    """
    wb = openpyxl.Workbook()

    # ----------------------------------------------------
    # Sheet 1: Summary & Metadata
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary & Metadata"
    ws_summary.views.sheetView[0].showGridLines = True

    # Disclaimer Header Banner (Row 1)
    ws_summary.merge_cells("A1:F1")
    cell_d = ws_summary["A1"]
    cell_d.value = "AI-Generated Draft — For Internal Engineering Review Only (Non-Certified Verdict)"
    cell_d.font = DISCLAIMER_FONT
    cell_d.alignment = Alignment(horizontal="center", vertical="center")
    cell_d.fill = ALERT_FILL
    ws_summary.row_dimensions[1].height = 24

    # Title (Row 3)
    ws_summary.merge_cells("A3:F3")
    cell_t = ws_summary["A3"]
    title_text = sanitize_cell_value(data.get("title", "TECHNICAL INSPECTION & COMPLIANCE SUMMARY REPORT"))
    cell_t.value = str(title_text)
    cell_t.font = TITLE_FONT
    cell_t.alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.row_dimensions[3].height = 28

    # Metadata Table (Rows 5-8)
    meta_fields = [
        ("Task ID / Ref:", sanitize_cell_value(data.get("task_id", "TASK-AUTONOMOUS-01"))),
        ("Facility / Unit:", sanitize_cell_value(data.get("facility", "Primary Refining Unit & Flare Header"))),
        ("Generated Timestamp (UTC):", sanitize_cell_value(data.get("timestamp", datetime.datetime.now(datetime.timezone.utc).isoformat()))),
        ("Evaluation Status:", sanitize_cell_value(data.get("status", "ACTION REQUIRED — NON-COMPLIANCE DETECTED"))),
    ]

    for idx, (label, val) in enumerate(meta_fields, start=5):
        ws_summary.cell(row=idx, column=1, value=label).font = BOLD_FONT
        ws_summary.cell(row=idx, column=1).fill = ICE_FILL
        ws_summary.cell(row=idx, column=1).border = THIN_BORDER
        ws_summary.cell(row=idx, column=1).alignment = Alignment(vertical="center")

        ws_summary.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=6)
        c_val = ws_summary.cell(row=idx, column=2, value=str(val))
        c_val.font = ALERT_FONT if "ACTION REQUIRED" in str(val) else REGULAR_FONT
        c_val.border = THIN_BORDER
        c_val.alignment = Alignment(vertical="center")
        for c in range(2, 7):
            ws_summary.cell(row=idx, column=c).border = THIN_BORDER
        ws_summary.row_dimensions[idx].height = 20

    # Executive Overview (Row 10)
    ws_summary.cell(row=10, column=1, value="Executive Inspection Overview").font = SECTION_FONT
    ws_summary.merge_cells("A11:F13")
    summary_text = sanitize_cell_value(data.get(
        "summary",
        "Autonomous document intelligence analysis executed on the submitted inspection report. "
        "Findings were benchmarked against sovereign indexed Standard Operating Procedures (SOPs), "
        "maintenance manuals, and equipment compliance thresholds."
    ))
    c_sum = ws_summary["A11"]
    c_sum.value = str(summary_text)
    c_sum.font = REGULAR_FONT
    c_sum.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c_sum.fill = ALT_ROW_FILL
    for r in range(11, 14):
        for c in range(1, 7):
            ws_summary.cell(row=r, column=c).border = THIN_BORDER

    # Footer Disclaimer (Row 15)
    ws_summary.merge_cells("A15:F15")
    ws_summary["A15"].value = "Confidential — Sovereign On-Premise AI Workbench — Non-Certified Draft"
    ws_summary["A15"].font = Font(name=FONT_NAME, size=9, italic=True, color="888888")
    ws_summary["A15"].alignment = Alignment(horizontal="center", vertical="center")

    auto_fit_columns(ws_summary)

    # ----------------------------------------------------
    # Sheet 2: Inspection Findings
    # ----------------------------------------------------
    ws_findings = wb.create_sheet(title="Inspection Findings")
    ws_findings.views.sheetView[0].showGridLines = True

    # Disclaimer Header
    ws_findings.merge_cells("A1:D1")
    ws_findings["A1"].value = "AI-Generated Draft — For Internal Engineering Review Only (Non-Certified Verdict)"
    ws_findings["A1"].font = DISCLAIMER_FONT
    ws_findings["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_findings["A1"].fill = ALERT_FILL

    # Section Title
    ws_findings.cell(row=3, column=1, value="Critical Inspection Findings & Anomalies").font = SECTION_FONT

    # Table Header
    headers_f = ["Finding #", "Observed Anomaly / Defect Description", "Severity Category", "Status"]
    for col_idx, h in enumerate(headers_f, start=1):
        c = ws_findings.cell(row=5, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = NAVY_FILL
        c.border = HEADER_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_findings.row_dimensions[5].height = 26

    raw_findings = data.get("critical_findings", [
        "Corrosion fatigue detected on primary discharge flange bolts exceeding 1.5mm wall thinning threshold.",
        "Pressure relief valve PRV-204 calibration interval exceeded maximum allowable 12-month limit.",
        "Emergency shutdown bypass valve seal integrity compromised with visible weeping of seal fluid."
    ])

    for row_idx, finding in enumerate(raw_findings, start=6):
        f_num = f"FND-{row_idx-5:02d}"
        f_desc = sanitize_cell_value(finding)
        severity = "CRITICAL" if "corrosion" in str(finding).lower() or "fatigue" in str(finding).lower() else "MAJOR"
        f_status = "Open / Action Required"

        c0 = ws_findings.cell(row=row_idx, column=1, value=f_num)
        c1 = ws_findings.cell(row=row_idx, column=2, value=str(f_desc))
        c2 = ws_findings.cell(row=row_idx, column=3, value=severity)
        c3 = ws_findings.cell(row=row_idx, column=4, value=f_status)

        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for cell in (c0, c1, c2, c3):
            cell.border = THIN_BORDER
            if fill.fill_type:
                cell.fill = fill
            cell.font = REGULAR_FONT
            cell.alignment = Alignment(vertical="center")

        c0.alignment = Alignment(horizontal="center", vertical="center")
        c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.font = ALERT_FONT if severity == "CRITICAL" else BOLD_FONT
        c3.alignment = Alignment(horizontal="center", vertical="center")
        ws_findings.row_dimensions[row_idx].height = 24

    auto_fit_columns(ws_findings, min_width=15, max_width=70)

    # ----------------------------------------------------
    # Sheet 3: Compliance & SOP Citations
    # ----------------------------------------------------
    ws_comp = wb.create_sheet(title="Compliance Gaps & Citations")
    ws_comp.views.sheetView[0].showGridLines = True

    # Disclaimer Header
    ws_comp.merge_cells("A1:D1")
    ws_comp["A1"].value = "AI-Generated Draft — For Internal Engineering Review Only (Non-Certified Verdict)"
    ws_comp["A1"].font = DISCLAIMER_FONT
    ws_comp["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_comp["A1"].fill = ALERT_FILL

    ws_comp.cell(row=3, column=1, value="Compliance Gaps & Authoritative Standard Citations").font = SECTION_FONT

    headers_c = ["Item #", "Observed Finding / Defect", "Authoritative SOP / Standard Citation", "Compliance Rating"]
    for col_idx, h in enumerate(headers_c, start=1):
        c = ws_comp.cell(row=5, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = NAVY_FILL
        c.border = HEADER_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_comp.row_dimensions[5].height = 26

    raw_gaps = data.get("compliance_gaps", [
        ("Discharge Flange Wall Thinning", "Safety SOP - Section 4.2 Emergency Shutdown Systems (p.12)", "CRITICAL NON-COMPLIANCE"),
        ("PRV-204 Recertification Overdue", "Equipment Standards - Section 11.4 Relief Valve Recertification (p.56)", "MAJOR GAP"),
        ("Seal Integrity Weeping", "Maintenance Manual - Section 8.1 Flange Integrity & Bolt Torquing (p.34)", "MODERATE GAP"),
    ])

    for row_idx, item in enumerate(raw_gaps, start=6):
        item_num = f"GAP-{row_idx-5:02d}"
        if isinstance(item, (list, tuple)) and len(item) >= 3:
            defect, citation, rating = sanitize_cell_value(item[0]), sanitize_cell_value(item[1]), sanitize_cell_value(item[2])
        else:
            defect, citation, rating = sanitize_cell_value(str(item)), "Authoritative Safety SOP", "NON-COMPLIANCE"

        c0 = ws_comp.cell(row=row_idx, column=1, value=item_num)
        c1 = ws_comp.cell(row=row_idx, column=2, value=str(defect))
        c2 = ws_comp.cell(row=row_idx, column=3, value=str(citation))
        c3 = ws_comp.cell(row=row_idx, column=4, value=str(rating))

        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for cell in (c0, c1, c2, c3):
            cell.border = THIN_BORDER
            if fill.fill_type:
                cell.fill = fill
            cell.font = REGULAR_FONT
            cell.alignment = Alignment(vertical="center")

        c0.alignment = Alignment(horizontal="center", vertical="center")
        c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c3.alignment = Alignment(horizontal="center", vertical="center")
        if "CRITICAL" in str(rating) or "NON-COMPLIANCE" in str(rating):
            c3.font = ALERT_FONT
        ws_comp.row_dimensions[row_idx].height = 24

    auto_fit_columns(ws_comp, min_width=15, max_width=60)

    # ----------------------------------------------------
    # Sheet 4: Recommendations
    # ----------------------------------------------------
    ws_recs = wb.create_sheet(title="Recommendations")
    ws_recs.views.sheetView[0].showGridLines = True

    # Disclaimer Header
    ws_recs.merge_cells("A1:D1")
    ws_recs["A1"].value = "AI-Generated Draft — For Internal Engineering Review Only (Non-Certified Verdict)"
    ws_recs["A1"].font = DISCLAIMER_FONT
    ws_recs["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_recs["A1"].fill = ALERT_FILL

    ws_recs.cell(row=3, column=1, value="Actionable Engineering Recommendations").font = SECTION_FONT

    headers_r = ["Priority", "Remediation Action Item", "Target Window", "Responsible Department"]
    for col_idx, h in enumerate(headers_r, start=1):
        c = ws_recs.cell(row=5, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = NAVY_FILL
        c.border = HEADER_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_recs.row_dimensions[5].height = 26

    raw_recs = data.get("recommendations", [
        "Initiate immediate scheduled depressurization and replacement of flange bolting assembly.",
        "Perform off-line hydrostatic bench testing and recertification for PRV-204 within 48 hours.",
        "Replace primary mechanical seal pack on pump P-102A prior to resuming continuous service."
    ])

    for row_idx, rec in enumerate(raw_recs, start=6):
        pri = f"P{row_idx-5}"
        action = sanitize_cell_value(rec)
        target = "48 Hours" if row_idx == 6 else ("7 Days" if row_idx == 7 else "14 Days")
        resp = "Mechanical Maintenance"

        c0 = ws_recs.cell(row=row_idx, column=1, value=pri)
        c1 = ws_recs.cell(row=row_idx, column=2, value=str(action))
        c2 = ws_recs.cell(row=row_idx, column=3, value=target)
        c3 = ws_recs.cell(row=row_idx, column=4, value=resp)

        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for cell in (c0, c1, c2, c3):
            cell.border = THIN_BORDER
            if fill.fill_type:
                cell.fill = fill
            cell.font = REGULAR_FONT
            cell.alignment = Alignment(vertical="center")

        c0.alignment = Alignment(horizontal="center", vertical="center")
        c0.font = ALERT_FONT if pri == "P1" else BOLD_FONT
        c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c3.alignment = Alignment(horizontal="center", vertical="center")
        ws_recs.row_dimensions[row_idx].height = 24

    auto_fit_columns(ws_recs, min_width=15, max_width=75)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
